from openpyxl import Workbook
wb = Workbook()

#Relative References
ws = wb.worksheets[0]
ws["A1"].value = 56596
ws["C3"].value =  "abc"

print(ws["A1"].value)
print(ws["C3"].value)


#Absolute References
ws.cell(row = 1, column = 1).value = 111
ws.cell(row = 3, column = 3).value = "def"
print(ws["A1"].value)
print(ws["C3"].value)

#To get column letter for certain index
from openpyxl.utils.cell import get_column_letter
print(get_column_letter(3))

#Output will be C

#Get column index from strings
from openpyxl.utils.cell import column_index_from_string
print(column_index_from_string("HH"))
