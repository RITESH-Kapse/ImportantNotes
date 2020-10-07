from openpyxl import Workbook
wb = Workbook()

ws = wb.worksheets[0]
ws["A1"].value = "Train"

ws.cell(1,1).offset(0,1).value = "Train cart"

print(ws["A1"].value , ws["B1"].value)

#That is new value will be added in B1 cell due to offset cammand above
