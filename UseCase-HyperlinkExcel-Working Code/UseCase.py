#Import the libraries
from openpyxl import load_workbook
from copy import copy
from openpyxl.utils import get_column_letter

"""Only Targetting first and last sheet"""
wb = load_workbook('Sample_Automation1.xlsx')
ws2 = wb["Sheet4"]
ws1 = wb["Sheet1"]


#Check the title of first sheet and last sheet
print(ws1.title)
print(ws2.title)


#Iterate over all the rows in sheet 1.
for row in ws1.iter_rows(min_row=4, max_col=ws1.max_column, max_row=ws1.max_row):
    for cell in row:
        if cell.value != None :
            print(cell.value)

            for row1 in ws2.iter_rows(min_row=6, max_col=ws2.max_column, max_row=ws2.max_row):
                for cell1 in row1:
                    # cell1_address = cell1.coordinate
                    if cell.value == cell1.value:
                        cell.hyperlink = '#%s!%s' % ("Sheet4", cell1.coordinate)
                        cell.style = "Hyperlink" 

#Just print seperation line as * to seperate each value in a for loop
    print("-"*50)  

"""Saving file as another name , since it can happend that another workbook is open """
wb.save("Sample_Automation1_updated.xlsx")
