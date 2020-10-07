from openpyxl import Workbook
wb = Workbook()

#Create the sheets
ws = wb.create_sheet("A sheet",0)
#If no index then it will go at last
w2 = wb.create_sheet("Sheet nr 2")

for sheet in wb:
    print(sheet.title)

#Remove the sheets
wb.remove(wb["Sheet"])
for sheet in wb:
    print(sheet.title)

#To remove particular sheet
del wb["A sheet"]
for sheet in wb:
    print(sheet.title)

#Save Workbook
wb.save("added_sheetss.xlsx")
