from openpyxl import load_workbook

wb = load_workbook("ComputerVisionFundamentals.xlsx")

for sheet in wb:
    print(sheet.title)

ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]

worksheets = wb.sheetnames
# type(worksheets)
#It will be list
print(worksheets)

ws3 = wb[worksheets[2]]
for sheet in wb:
    print(sheet.title)

print(ws3.title)
